#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从机台 SQLite 导出游戏记录风格表（对照 FIREBIRD 样例 xlsx 的 T 表前 10 列）。

列映射：
  A 大奖统计     -> 赢/输/免费/大奖/彩金；「大奖」= result_type 为 BonusWin(3)；
                 「彩金」= jackpot_win&gt;0 或 result_type 为 Jackpot/JackpotOnline（与 Unity 一致）
                 可选 --col-a id / row_index
  B–J            -> 下注前/下注/各分项得分/总得分/输赢/结束；免费合并段「下注」仅触发局 total_bet

依赖：Python 3.8+；写 xlsx 需 openpyxl（pip install openpyxl）。
"""

from __future__ import annotations

import argparse
import csv
import sqlite3
import sys
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence, Tuple

OPEN_NORMAL = 0
OPEN_GIVE = 1
RT_FREE_WIN = 2
RT_BONUS_WIN = 3
RT_JACKPOT = 4
RT_JACKPOT_ONLINE = 5

HEADER_ROW: Tuple[str, ...] = (
    "大奖统计",
    "下注前",
    "下注",
    "基础游戏得分",
    "免费游戏得分",
    "大奖游戏得分",
    "彩金得分",
    "总得分",
    "输赢",
    "结束",
)


def connect(db_path: Path) -> sqlite3.Connection:
    if not db_path.is_file():
        raise FileNotFoundError(f"找不到 SQLite 文件: {db_path}")
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def fetch_rows(
    conn: sqlite3.Connection,
    game_id: int,
    limit: Optional[int],
    order: str,
) -> List[sqlite3.Row]:
    order_sql = "ASC" if order.lower() == "asc" else "DESC"
    lim_sql = f" LIMIT {int(limit)} " if limit is not None and limit > 0 else ""
    sql = f"""
        SELECT
            id,
            game_id,
            total_bet,
            credit_before,
            credit_after,
            base_game_win_credit,
            free_spin_win_credit,
            bonus_game_win_credit,
            jackpot_win_credit,
            total_win_credit,
            open_type,
            result_type,
            free_curtime,
            free_totaltime
        FROM slot_game_record
        WHERE game_id = ?
        ORDER BY id {order_sql}
        {lim_sql}
    """
    cur = conn.execute(sql, (game_id,))
    return list(cur.fetchall())


def _iget(r: sqlite3.Row, key: str) -> int:
    v = r[key]
    if v is None:
        return 0
    return int(v)


def _lget(r: sqlite3.Row, key: str) -> int:
    v = r[key]
    if v is None:
        return 0
    return int(v)


def scale(v: Any, divisor: float) -> float:
    if v is None:
        return 0.0
    return float(v) / divisor


def is_free_trigger_row(r: sqlite3.Row) -> bool:
    return (
        _iget(r, "open_type") == OPEN_NORMAL
        and _iget(r, "result_type") == RT_FREE_WIN
        and _iget(r, "free_totaltime") > 0
    )


def is_jackpot_prize_row(r: sqlite3.Row) -> bool:
    """彩金：有 jackpot_win 或 result_type 为 Jackpot / JackpotOnline（不含 BonusWin）。"""
    if _lget(r, "jackpot_win_credit") > 0:
        return True
    rt = _iget(r, "result_type")
    return rt in (RT_JACKPOT, RT_JACKPOT_ONLINE)


def classify_merged_segment(seg: Sequence[sqlite3.Row]) -> str:
    for r in seg:
        if is_jackpot_prize_row(r):
            return "彩金"
    return "免费"


def classify_win_lose(r: sqlite3.Row, divisor: float) -> str:
    """主游戏：基础游戏得分为 0 为输，否则为赢。"""
    _ = divisor
    return "赢" if (_lget(r, "base_game_win_credit") or 0) > 0 else "输"


def classify_single_row(r: sqlite3.Row, divisor: float, big_mul: float) -> str:
    _ = big_mul
    if is_jackpot_prize_row(r):
        return "彩金"
    if _iget(r, "result_type") == RT_BONUS_WIN:
        return "大奖"
    if _iget(r, "open_type") == OPEN_GIVE or _iget(r, "result_type") == RT_FREE_WIN:
        return "免费"
    return classify_win_lose(r, divisor)


def build_merged_segments(rows: Sequence[sqlite3.Row]) -> List[List[sqlite3.Row]]:
    segs: List[List[sqlite3.Row]] = []
    i = 0
    n = len(rows)
    while i < n:
        r = rows[i]
        if is_free_trigger_row(r):
            seg = [r]
            j = i + 1
            while j < n and _iget(rows[j], "open_type") == OPEN_GIVE:
                seg.append(rows[j])
                j += 1
            segs.append(seg)
            i = j
        elif _iget(r, "open_type") == OPEN_GIVE:
            seg = []
            while i < n and _iget(rows[i], "open_type") == OPEN_GIVE:
                seg.append(rows[i])
                i += 1
            segs.append(seg)
        else:
            segs.append([r])
            i += 1
    return segs


def format_col_a(
    mode: str,
    label: str,
    seg: Sequence[sqlite3.Row],
    line_no: int,
) -> Any:
    if mode == "id":
        return int(seg[0]["id"])
    if mode == "row_index":
        return line_no
    return label


def segment_to_row(
    seg: Sequence[sqlite3.Row],
    divisor: float,
    col_a_mode: str,
    line_no: int,
    big_mul: float,
) -> Tuple[Any, ...]:
    label = classify_merged_segment(seg) if len(seg) > 1 else classify_single_row(seg[0], divisor, big_mul)
    col_a = format_col_a(col_a_mode, label, seg, line_no)

    first = seg[0]
    # 「免费」合并段：下注列只用触发局 total_bet，不累加赠送局
    if len(seg) > 1 and label == "免费":
        sum_bet = _lget(first, "total_bet")
    else:
        sum_bet = sum(_lget(r, "total_bet") for r in seg)
    sum_base = sum(_lget(r, "base_game_win_credit") or 0 for r in seg)
    sum_free_raw = sum(_lget(r, "free_spin_win_credit") or 0 for r in seg)
    sum_bonus = sum(_lget(r, "bonus_game_win_credit") or 0 for r in seg)
    sum_jp = sum(_lget(r, "jackpot_win_credit") or 0 for r in seg)
    sum_total = sum(_lget(r, "total_win_credit") or 0 for r in seg)
    if sum_total == 0 and (sum_base + sum_free_raw + sum_bonus + sum_jp) != 0:
        sum_total = sum_base + sum_free_raw + sum_bonus + sum_jp
    last = seg[-1]

    total_bet = scale(sum_bet, divisor)
    cb = scale(first["credit_before"], divisor)
    ca = scale(last["credit_after"], divisor)
    sum_free = sum_free_raw
    if len(seg) > 1 and label == "免费":
        canonical = _lget(last, "credit_after") - _lget(first, "credit_before") + sum_bet
        if canonical < 0:
            canonical = 0
        sum_total = canonical
        sum_free = canonical - sum_base - sum_bonus - sum_jp
        if sum_free < 0:
            sum_free = 0

    base_g = scale(sum_base, divisor)
    free_g = scale(sum_free, divisor)
    bonus_g = scale(sum_bonus, divisor)
    jp = scale(sum_jp, divisor)
    total_w = scale(sum_total, divisor)
    delta = ca - cb
    line_no_jp = base_g + free_g + bonus_g
    delta_parts = line_no_jp + jp - total_bet
    if abs(delta - delta_parts) > max(1e-6, 1e-4 * max(abs(delta), 1.0)):
        print(
            f"[WARN] id={first['id']} delta={delta:.6g} != 基础+免费+大奖+彩金-下注={delta_parts:.6g}",
            file=sys.stderr,
        )
    return (
        col_a,
        cb,
        total_bet,
        base_g,
        free_g,
        bonus_g,
        None if jp == 0 else jp,
        total_w,
        delta,
        ca,
    )


def build_all_rows(
    rows: Sequence[sqlite3.Row],
    divisor: float,
    col_a_mode: str,
    merge: bool,
    big_mul: float,
) -> List[Tuple[Any, ...]]:
    segs = build_merged_segments(rows) if merge else [[r] for r in rows]
    out: List[Tuple[Any, ...]] = []
    for i, seg in enumerate(segs, start=1):
        out.append(segment_to_row(seg, divisor, col_a_mode, i, big_mul))
    return out


def write_csv(path: Path, rows: Iterable[Tuple[Any, ...]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADER_ROW)
        for tup in rows:
            w.writerow(list(tup))


def write_xlsx(
    path: Path,
    rows: List[Tuple[Any, ...]],
    sheet_name: str,
    template_path: Optional[Path],
) -> None:
    try:
        from openpyxl import Workbook as OpenpyxlWorkbook
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("写 xlsx 需要 openpyxl：pip install openpyxl") from e

    path.parent.mkdir(parents=True, exist_ok=True)

    if template_path and template_path.is_file():
        wb = load_workbook(template_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        tpl = wb["T"] if "T" in wb.sheetnames else None
        if tpl is not None:
            for c in range(1, len(HEADER_ROW) + 2):
                v = tpl.cell(row=1, column=c).value
                ws.cell(row=1, column=c, value=v)
        else:
            for i, h in enumerate(HEADER_ROW, start=1):
                ws.cell(row=1, column=i, value=h or None)
        start_row = 2
    else:
        wb = OpenpyxlWorkbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        for i, h in enumerate(HEADER_ROW, start=1):
            ws.cell(row=1, column=i, value=h or None)
        start_row = 2

    for offset, tup in enumerate(rows, start=0):
        for col, val in enumerate(tup, start=1):
            ws.cell(row=start_row + offset, column=col, value=val)
    wb.save(path)


def print_preview(conn: sqlite3.Connection, game_id: int, n: int) -> None:
    rows = fetch_rows(conn, game_id, limit=n, order="asc")
    if not rows:
        print(f"(无数据) game_id={game_id}")
        return
    cols = [
        "id",
        "open",
        "rt",
        "free_t",
        "total_bet",
        "credit_before",
        "credit_after",
        "win",
        "jp",
        "delta",
    ]
    print("\t".join(cols))
    for r in rows:
        cb = r["credit_before"]
        ca = r["credit_after"]
        tb = r["total_bet"]
        win = r["base_game_win_credit"] or 0
        jp = r["jackpot_win_credit"] or 0
        delta = ca - cb
        print(
            "\t".join(
                str(x)
                for x in (
                    r["id"],
                    r["open_type"],
                    r["result_type"],
                    r["free_totaltime"],
                    tb,
                    cb,
                    ca,
                    win,
                    jp,
                    delta,
                )
            )
        )


def main() -> None:
    ap = argparse.ArgumentParser(description="slot_game_record -> 游戏记录 CSV/xlsx")
    ap.add_argument(
        "--db",
        type=Path,
        required=True,
        help="SQLite 路径（Editor 常为项目 Assets/StreamingAssets/<dbName>）",
    )
    ap.add_argument("--game-id", type=int, default=3997)
    ap.add_argument("--limit", type=int, default=0, help="0 表示导出全部")
    ap.add_argument("--order", choices=("asc", "desc"), default="asc")
    ap.add_argument(
        "--credit-divisor",
        type=float,
        default=1.0,
        help="若库内为整数分、Excel 要元，填 100",
    )
    ap.add_argument("--preview", type=int, default=0, help="仅打印前 N 条对照，不写文件")
    ap.add_argument("--out-csv", type=Path, default=None)
    ap.add_argument("--out-xlsx", type=Path, default=None)
    ap.add_argument(
        "--template",
        type=Path,
        default=None,
        help="可选：参考 xlsx，用于复制第 1 行表头/说明（从工作表 T 读第 1 行）",
    )
    ap.add_argument(
        "--sheet-name",
        default="3997",
        help="写入 xlsx 的工作表名（与 --template 合用时为新增表）",
    )
    ap.add_argument(
        "--col-a",
        choices=("label", "id", "row_index"),
        default="label",
        help="A 列：局类型文案 label（默认）/ id / row_index",
    )
    ap.add_argument(
        "--no-merge-free",
        action="store_true",
        help="不合并免费触发+赠送局（每库行一行）",
    )
    ap.add_argument(
        "--big-win-multiple",
        type=float,
        default=15.0,
        help="已保留兼容；「大奖」行由 result_type=BonusWin(3) 判定，不再用线赢倍数阈值",
    )
    args = ap.parse_args()

    conn = connect(args.db)
    try:
        if args.preview > 0:
            print_preview(conn, args.game_id, args.preview)
            return

        limit = None if args.limit <= 0 else args.limit
        raw = fetch_rows(conn, args.game_id, limit=limit, order=args.order)
        if not raw:
            print(f"(无数据) game_id={args.game_id}", file=sys.stderr)
            return

        merge = not args.no_merge_free
        out_rows = build_all_rows(
            raw,
            args.credit_divisor,
            args.col_a,
            merge,
            args.big_win_multiple,
        )

        if args.out_csv:
            write_csv(args.out_csv, out_rows)
            print(f"Wrote CSV: {args.out_csv.resolve()} ({len(out_rows)} 行, 库 {len(raw)} 条)")
        if args.out_xlsx:
            write_xlsx(
                args.out_xlsx,
                out_rows,
                sheet_name=args.sheet_name[:31],
                template_path=args.template,
            )
            print(f"Wrote XLSX: {args.out_xlsx.resolve()}")

        if not args.out_csv and not args.out_xlsx:
            print("请指定 --out-csv 或 --out-xlsx，或使用 --preview", file=sys.stderr)
            sys.exit(2)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
